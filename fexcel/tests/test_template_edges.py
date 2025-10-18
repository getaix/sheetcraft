import os
import json
import tempfile

import jinja2
import pytest

from fexcel.template import ExcelTemplate


def test_template_init_image_extension_failure(monkeypatch):
    # 让 add_extension 抛错以覆盖 __init__ 中的异常吞掉分支
    def boom(self, ext):
        raise RuntimeError("boom")

    monkeypatch.setattr(jinja2.Environment, "add_extension", boom)
    t = ExcelTemplate()
    # 仍应成功初始化环境
    assert t._env is not None


def test_template_filters_update_and_render():
    t = ExcelTemplate(filters={"my_upper": lambda s: s.upper()})
    out = t._render_string("{{ 'x'|my_upper }}", {})
    assert out == "X"


def test_template_render_in_cell_keep_ratio_pil_fail():
    # 构造包含图片占位符的简单模板，路径指向不存在的文件，触发 keep_ratio 的异常分支
    fd, tpl_path = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    fd2, out_path = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd2)
    try:
        from openpyxl import Workbook, load_workbook

        wb = Workbook()
        ws = wb.active
        payload = {
            "path": "no_such_file.png",
            "in_cell": True,
            "keep_ratio": True,
        }
        ws.cell(row=1, column=1).value = "__FEXCEL_IMG__" + json.dumps(payload)
        wb.save(tpl_path)

        t = ExcelTemplate()
        t.render(tpl_path, {}, out_path)

        out_wb = load_workbook(out_path)
        out_ws = out_wb.active
        # 处理后应清空占位单元格，不抛异常
        assert out_ws.cell(row=1, column=1).value is None
    finally:
        for p in (tpl_path, out_path):
            try:
                os.remove(p)
            except OSError:
                pass